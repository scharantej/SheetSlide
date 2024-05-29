
from flask import Flask, render_template, request, redirect, url_for, flash
from werkzeug.utils import secure_filename
import openpyxl

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/create_slide', methods=['GET', 'POST'])
def create_slide():
    if request.method == 'POST':
        title = request.form['title']
        content = request.form['content']
        file = request.files['file']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            data = []
            for row in sheet.iter_rows():
                data.append([cell.value for cell in row])
            flash('Slide created successfully!')
            return redirect(url_for('view_slides'))
    return render_template('create_slide.html')

@app.route('/view_slides')
def view_slides():
    slides = os.listdir(app.config['UPLOAD_FOLDER'])
    return render_template('view_slides.html', slides=slides)

@app.route('/view_slide/<slide_id>')
def view_slide(slide_id):
    slide = os.path.join(app.config['UPLOAD_FOLDER'], slide_id)
    wb = openpyxl.load_workbook(slide)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    return render_template('view_slide.html', data=data)

@app.route('/edit_slide/<slide_id>', methods=['GET', 'POST'])
def edit_slide(slide_id):
    slide = os.path.join(app.config['UPLOAD_FOLDER'], slide_id)
    if request.method == 'POST':
        title = request.form['title']
        content = request.form['content']
        wb = openpyxl.load_workbook(slide)
        sheet = wb.active
        sheet['A1'] = title
        sheet['B1'] = content
        wb.save(slide)
        flash('Slide updated successfully!')
        return redirect(url_for('view_slides'))
    wb = openpyxl.load_workbook(slide)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    return render_template('edit_slide.html', data=data)

if __name__ == '__main__':
    app.secret_key = 'supersecretkey'
    app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'slides')
    app.run(debug=True)
